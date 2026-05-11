"""
Export an EGalvanicWorkbook to a properly formatted .xlsx file.

Follows the K STAR / eGalvanic schema:
- 18 columns (15 base + 3 optional)
- Frozen header row, bold + gray header style
- ID columns left blank (eGalvanic generates them on upload)
"""
from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.models.schema import EGalvanicWorkbook

logger = logging.getLogger(__name__)


ASSET_COLUMNS = [
    'asset_id', 'asset_name', 'asset_class', 'asset_subtype', 'building',
    'floor', 'room', 'com', 'criticality', 'operating_conditions',
    'maintenance_state', 'suggested_shortcut', 'qr_code', 'parent_asset_name',
    'delete', 'voltage', 'ampere_rating', 'kva_rating',
]

CONNECTION_COLUMNS = [
    'connection_id', 'source_asset_name', 'source_handle',
    'target_asset_name', 'target_handle', 'connection_class', 'delete',
]


def write_workbook(workbook: EGalvanicWorkbook, out_path: str | Path) -> Path:
    """Write the workbook to an .xlsx file."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # --- Assets sheet ---
    ws_a = wb.active
    ws_a.title = 'Assets'
    ws_a.append(ASSET_COLUMNS)
    for asset in workbook.assets:
        ws_a.append([
            asset.asset_id, asset.asset_name, asset.asset_class, asset.asset_subtype,
            asset.building, asset.floor, asset.room, asset.com,
            asset.criticality, asset.operating_conditions, asset.maintenance_state,
            asset.suggested_shortcut, asset.qr_code, asset.parent_asset_name,
            asset.delete, asset.voltage, asset.ampere_rating, asset.kva_rating,
        ])

    # --- Connections sheet ---
    ws_c = wb.create_sheet('Connections')
    ws_c.append(CONNECTION_COLUMNS)
    for conn in workbook.connections:
        ws_c.append([
            conn.connection_id, conn.source_asset_name, conn.source_handle,
            conn.target_asset_name, conn.target_handle, conn.connection_class,
            conn.delete,
        ])

    # Styling
    for ws in (ws_a, ws_c):
        for cell in ws[1]:
            cell.font = Font(bold=True, name='Arial', size=11)
            cell.fill = PatternFill('solid', start_color='D9D9D9')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.freeze_panes = 'A2'
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name='Arial', size=10)
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 22

    wb.save(out_path)
    logger.info(f"Wrote {len(workbook.assets)} assets and {len(workbook.connections)} connections to {out_path}")
    return out_path
